from datetime import datetime
from django.db.models import Q
from django.utils import timezone
from decimal import Decimal
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.views.generic import ListView
from .models import Product, Sale, Profit
from .forms import ProductForm, SaleForm, ProfitForm,Profit,PaymentUpdateForm
from django.shortcuts import redirect
from django.db.models import  F,  Sum
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from excel_response import ExcelResponse
from django.db.models import Max
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Border, Side



def index(request):
    product_list = Product.objects.all()
    search_query = request.GET.get('search', '').lower()
    filtered_products = product_list.filter(Q(nomi__icontains=search_query))

    paginator_products = Paginator(filtered_products, 15)
    page_products = request.GET.get('page')
    try:
        products = paginator_products.page(page_products)
    except PageNotAnInteger:
        products = paginator_products.page(1)
    except EmptyPage:
        products = paginator_products.page(paginator_products.num_pages)

    latest_sales = Sale.objects.values('dokon').annotate(last_sold=Max('sotilgan_sana')).order_by()

    # Формируем список последних проданных товаров
    latest_sales_list = []
    for sale in latest_sales:
        latest_sale = Sale.objects.filter(dokon=sale['dokon'], sotilgan_sana=sale['last_sold']).first()
        if latest_sale:
            latest_sales_list.append(latest_sale)

    date_str = request.GET.get('date')

    if date_str:
        date = datetime.strptime(date_str, '%Y-%m-%d').date()
        filtered_sales = sales_list.filter(sotilgan_sana=date)
    else:
        filtered_sales = latest_sales_list

    paginator_sales = Paginator(filtered_sales, 15)
    page_sales = request.GET.get('page_sales')
    try:
        sales = paginator_sales.page(page_sales)
    except PageNotAnInteger:
        sales = paginator_sales.page(1)
    except EmptyPage:
        sales = paginator_sales.page(paginator_sales.num_pages)

    total_balance = Product.objects.aggregate(total_balance=Sum(F('soni') * F('tan_narxi')))['total_balance'] or 0
    total_amount = Profit.objects.aggregate(olinadigan_summa=Sum('olinadigan_summa'))['olinadigan_summa'] or 0

    latest_profit = Profit.objects.last()
    if latest_profit is None:
        latest_profit = Profit()
        latest_profit.save()
    total_profit = latest_profit.total_profit

    total_balance_formatted = "${:,.2f}".format(total_balance)
    total_profit_formatted = "${:,.2f}".format(total_profit)
    total_amount_formatted = "${:,.2f}".format(total_amount)

    if 'export_products_excel' in request.GET:
        try:
            # Создаем новую книгу Excel
            wb = Workbook()
            ws = wb.active

            # Добавляем заголовки в первую строку
            ws.append(['Mahsulot', 'Narxi', 'Soni'])

            # Получаем все продукты из базы данных
            products = Product.objects.all()

            # Добавляем данные о продуктах в книгу Excel
            for product in products:
                ws.append([product.nomi, product.tan_narxi, product.soni])

            # Добавляем границы к ячейкам таблицы
            for row in ws.iter_rows(min_row=1, max_row=len(products) + 1, min_col=1, max_col=3):
                for cell in row:
                    cell.border = Border(left=Side(border_style='thin'),
                                         right=Side(border_style='thin'),
                                         top=Side(border_style='thin'),
                                         bottom=Side(border_style='thin'))

            # Создаем объект BytesIO для временного хранения данных Excel в памяти
            excel_file = BytesIO()

            # Сохраняем книгу Excel в объект BytesIO
            wb.save(excel_file)

            # Устанавливаем указатель файла в начало
            excel_file.seek(0)

            # Создаем HTTP-ответ с содержимым Excel-файла в виде прикрепленного файла
            response = HttpResponse(excel_file.getvalue(),
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="products.xlsx"'

            # Возвращаем HTTP-ответ
            return response

        except ValueError:
            pass



    context = {
        'products': products,
        'latest_sales_list': latest_sales_list,
        'total_balance': total_balance_formatted,
        'total_profit': total_profit_formatted,
        'total_amount': total_amount_formatted,
        'search_query': search_query,
    }

    return render(request, 'index.html', context)



def add_product(request):
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            try:
                nomi = form.cleaned_data['nomi']
                existing_product = Product.objects.filter(nomi=nomi).first()
                if existing_product:
                    existing_product.soni += form.cleaned_data['soni']
                    existing_product.tan_narxi = form.cleaned_data['tan_narxi']
                    existing_product.save()
                else:
                    form.save()

                return redirect('index')
            except Exception as e:
                print(f"Error adding product: {e}")
                return HttpResponse("Error adding product. Please try again.")
    else:
        form = ProductForm()

    return render(request, 'add_product.html', {'form': form, })


def sell_product(request):
    if request.method == 'POST':
        form = SaleForm(request.POST)
        if form.is_valid():
            sale = form.save(commit=False)
            product = form.cleaned_data['mahsulot']
            sale.name = product.nomi
            sale.foyda = sale.sotiladigan_narxi * sale.soni - product.tan_narxi * sale.soni

            if sale.tolash_usuli == 'Nasiya':
                sale.qarz = sale.sotiladigan_narxi * sale.soni
            else:
                sale.qarz = 0

            sale.save()

            sale.soni -= product.soni
            product.save()

            return redirect('index')
    else:
        form = SaleForm()

    context = {
        'form': form,
    }
    return render(request, 'sell_product.html', context)


def update_payment(request, sale_id):
    sale = get_object_or_404(Sale, pk=sale_id)
    form = PaymentUpdateForm(request.POST or None)

    if request.method == 'POST':
        if form.is_valid():
            amount_paid = form.cleaned_data['amount_paid']
            payment_update = form.save(commit=False)
            payment_update.sale = sale
            payment_update.save()

            sale.qarz -= amount_paid

            if sale.qarz <= 0:
                sale.tolash_usuli = 'Naqd'

            sale.save()

            return redirect('index')

    # Получить все платежи, связанные с текущей продажей
    payments = sale.paymentupdate_set.all()


    context = {
        'sale': sale,
        'form': form,
        'payments': payments,  # Добавляем список платежей в контекст
    }
    return render(request, 'update_payment.html', context)


def edit_sale(request, sale_id):
    sale = get_object_or_404(Sale, pk=sale_id)
    old_quantity = sale.soni
    product = sale.mahsulot

    if request.method == 'POST':
        form = SaleForm(request.POST, instance=sale)
        if form.is_valid():
            edited_sale = form.save(commit=False)

            new_quantity = edited_sale.soni
            difference = new_quantity - old_quantity  # Отнимаем новое значение от старого

            # Обновляем количество товара
            product.soni -= difference
            product.save()

            # Вычисляем прибыль
            edited_sale.foyda = (edited_sale.sotiladigan_narxi * edited_sale.soni) - (product.tan_narxi * edited_sale.soni)

            # Добавляем к долгу, если выбран способ оплаты "Nasiya"
            if edited_sale.tolash_usuli == 'Nasiya':
                edited_sale.qarz += edited_sale.sotiladigan_narxi * difference
            #     2600 110 10

            edited_sale.save()

            return redirect('index')
    else:
        form = SaleForm(instance=sale)

    context = {
        'form': form,
        'sale': sale,
    }
    return render(request, 'edit_sale.html', context)




def take_money(request):
    if request.method == 'POST':
        # Обработка POST-запроса для формы прибыли
        form = ProfitForm(request.POST)
        if form.is_valid():
            olinadigan_summa = form.cleaned_data['olinadigan_summa']
            kommentariya = form.cleaned_data['kommentariya']

            if olinadigan_summa is not None:
                total_sale_profit = Sale.objects.aggregate(foyda=Sum('foyda'))['foyda'] or 0
                latest_profit = Profit.objects.last()

                if latest_profit is not None:
                    latest_profit.total_profit -= olinadigan_summa
                    latest_profit.save()
                    profit = Profit.objects.create(olinadigan_summa=olinadigan_summa, sana=timezone.now(), kommentariya=kommentariya)
                    profit.total_profit = latest_profit.total_profit
                    profit.save()
                else:
                    profit = Profit.objects.create(olinadigan_summa=olinadigan_summa, sana=timezone.now(), kommentariya=kommentariya)
                    profit.total_profit = total_sale_profit - olinadigan_summa
                    profit.save()
                return redirect('index')
            else:
                return HttpResponse("Amount cannot be None.")
    else:
        form = ProfitForm()

    profits = Profit.objects.all()

    today = timezone.now().date()
    sales_list = Sale.objects.filter(sotilgan_sana=today).order_by('-sotilgan_sana')

    if 'export_sales_excel' in request.GET:
        try:
            wb = Workbook()
            ws = wb.active
            ws.append(['Mahsulot', 'Sotilgan narxi', 'Soni', 'Sana', 'Do`kon', 'Tolov usuli', 'Qarz', 'Foyda'])

            for sale in sales_list:
                formatted_date = sale.sotilgan_sana.strftime('%Y-%m-%d')
                ws.append([sale.mahsulot.nomi, sale.sotiladigan_narxi, sale.soni, formatted_date, sale.dokon,
                           sale.tolash_usuli, sale.qarz, sale.foyda])

            for row in ws.iter_rows(min_row=1, max_row=len(sales_list) + 1, min_col=1, max_col=8):
                for cell in row:
                    cell.border = Border(left=Side(border_style='thin'),
                                         right=Side(border_style='thin'),
                                         top=Side(border_style='thin'),
                                         bottom=Side(border_style='thin'))

            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)

            response = HttpResponse(excel_file.getvalue(),
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="sales.xlsx"'

            return response

        except ValueError:
            pass

    paginator = Paginator(sales_list, 10)
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    context = {
        'form': form,
        'profits': profits,
        'sales_list': sales_list,
    }

    return render(request, 'take_money.html', context)

class SaleListView(ListView):
    model = Sale
    template_name = 'sale_list.html'
    context_object_name = 'sales'
    paginate_by = 5

    def get(self, request, dokon, *args, **kwargs):
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        all_records = request.GET.get('all')

        if 'export_sales_excel' in request.GET:
            try:
                if all_records and all_records.lower() == 'true':
                    sales_queryset = Sale.objects.filter(dokon=dokon)

                    filename = f'{dokon} hamma sotilgan mahsuloti.xlsx'
                else:
                    if start_date and end_date:
                        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                        sales_queryset = Sale.objects.filter(sotilgan_sana__range=(start_date, end_date), dokon=dokon)
                    else:
                        sales_queryset = Sale.objects.filter(dokon=dokon)
                    filename = f'{dokon}sotilgan mahsuloti.xlsx'

                wb = Workbook()
                ws = wb.active
                ws.append(['Nomi', 'Sotilgan narxi', 'Soni', 'Sana', 'Do`kon', 'Tolov usuli', 'Qarz', 'Foyda'])
                for sale in sales_queryset:
                    formatted_date = sale.sotilgan_sana.strftime('%Y-%m-%d')
                    ws.append([sale.mahsulot.nomi, sale.sotiladigan_narxi, sale.soni, formatted_date, sale.dokon,
                               sale.tolash_usuli, sale.qarz, sale.foyda])

                for row in ws.iter_rows(min_row=1, max_row=len(sales_queryset) + 1, min_col=1, max_col=8):
                    for cell in row:
                        cell.border = Border(left=Side(border_style='thin'),
                                             right=Side(border_style='thin'),
                                             top=Side(border_style='thin'),
                                             bottom=Side(border_style='thin'))

                excel_file = BytesIO()
                wb.save(excel_file)
                excel_file.seek(0)

                response = HttpResponse(excel_file.getvalue(),
                                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="{filename}"'

                return response

            except ValueError:
                pass

        total_qarz = Sale.objects.filter(dokon=dokon).aggregate(total_qarz=Sum('qarz'))['total_qarz'] or 0
        sales = self.get_queryset().filter(dokon=dokon)
        paginator = Paginator(sales, self.paginate_by)
        page_number = request.GET.get('page')

        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

        form = SaleForm(initial={'dokon': dokon})
        context = {
            'form': form,
            'dokon': dokon,
            'page_obj': page_obj,
            'sales': page_obj,
            'total_qarz': total_qarz,
            'start_date': start_date,
            'end_date': end_date,
        }
        return render(request, self.template_name, context)

    def post(self, request, dokon, *args, **kwargs):
        form = SaleForm(request.POST)
        if form.is_valid():
            sale = form.save(commit=False)
            product = form.cleaned_data['mahsulot']
            sale.name = product.nomi
            sale.foyda = sale.sotiladigan_narxi * sale.soni - product.tan_narxi * sale.soni

            if sale.tolash_usuli == 'Nasiya':
                sale.qarz = sale.sotiladigan_narxi * sale.soni
            else:
                sale.qarz = 0

            sale.dokon = dokon
            sale.save()

            sale.soni -= product.soni
            product.save()

            return redirect('sale_list', dokon=dokon)


        else:
            total_qarz = Sale.objects.filter(dokon=dokon).aggregate(total_qarz=Sum('qarz'))['total_qarz'] or 0
            sales = self.get_queryset().filter(dokon=dokon)
            paginator = Paginator(sales, self.paginate_by)
            page_number = request.GET.get('page')

            try:
                page_obj = paginator.page(page_number)
            except PageNotAnInteger:
                page_obj = paginator.page(1)
            except EmptyPage:
                page_obj = paginator.page(paginator.num_pages)

            return render(request, self.template_name,
                          {'form': form, 'dokon': dokon, 'page_obj': page_obj, 'sales': page_obj,
                           'total_qarz': total_qarz})

    def get_queryset(self):
        dokon = self.kwargs.get('dokon')
        return Sale.objects.filter(dokon=dokon).order_by('-sotilgan_sana')